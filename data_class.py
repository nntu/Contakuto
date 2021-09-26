from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
import contakuto_basegui
import openpyxl as xl

class UserProfileData:
    def __init__(self, username):
        # Load data forom file
        self.username = username
        self.filepath = "./data/{0}/{0}.xlsx".format(self.username)
        self.workbook = xl.load_workbook(self.filepath)
        self.worksheet = self.workbook["Sheet"]
        self.data = {
            "fullname":         self.worksheet["A2"],
            "username":         self.worksheet["B2"],
            "password":         self.worksheet["C2"],
            "pin":              self.worksheet["D2"],
            "language":         self.worksheet["E2"],
            "picture":          self.worksheet["F2"],
            "theme":            self.worksheet["G2"]
        }

    def getData(self, field):
        return self.data[field].value

    def setData(self, field, new_data):
        self.data[field].value = new_data

    # Create list of contacts from data file
    def setupContactList(self, list):
        self.picPath = "./data/{0}/profile_pictures/".format(self.username)
        self.current_row = 6
        for row in self.worksheet.iter_rows(min_row=6, min_col=1, max_col=1, values_only=False):
            for cell in row:
                # Check if cell isn't empty
                if cell.value:
                    # Create custom listWidget
                    customListWidget = contakuto_basegui.QListWidgetX()
                    # Get value
                    nicknameValue = self.worksheet.cell(row=cell.row, column=2).value
                    companyValue = self.worksheet.cell(row=cell.row, column=3).value
                    picValue = self.worksheet.cell(row=cell.row, column=5).value
                    itemID = self.worksheet.cell(row=cell.row, column=20).value
                    # Set item icon
                    if not picValue:
                        iconValue = "./image/default_avatar.png"
                    else:
                        iconValue = self.picPath + picValue
                    iconPixmap = QPixmap(iconValue)
                    iconPixmap = iconPixmap.scaled(
                        40, 40,
                        aspectRatioMode=Qt.IgnoreAspectRatio, transformMode=Qt.SmoothTransformation)
                    customListWidget.setID(itemID)
                    customListWidget.setTextUp(cell.value)
                    if not nicknameValue and not companyValue:
                        pass
                    elif nicknameValue and not companyValue:
                        customListWidget.setTextDown(nicknameValue)
                    elif not nicknameValue and companyValue:
                        customListWidget.setTextDown(companyValue)
                    else:
                        customListWidget.setTextDown(nicknameValue + ", " + companyValue)
                    customListWidget.setIcon(iconPixmap)
                    # Add cell value to list
                    # Create item
                    item = QListWidgetItem(cell.value)
                    item.setSizeHint(customListWidget.sizeHint())
                    list.addItem(item)
                    list.setItemWidget(item, customListWidget)
                    #list.sortItems(Qt.AscendingOrder)
                self.current_row += 1

    def getEmptyRow(self):
        self.emptycell_counter = 0
        self.row = 6
        # Check empty row to find the last row which has value
        for row in self.worksheet.iter_rows(min_row=6, min_col=1, max_col=40):
            # If an empty row is found, break the loop
            if self.emptycell_counter == 40:
                break
            # If not an empty row
            else:
                for cell in row:
                    # If a cell is empty, increase counter value
                    if not cell.value:
                        self.emptycell_counter += 1
                    else:
                        self.emptycell_coutner = 0
            self.row += 1
        return self.row

    def getContactData(self, contact_row, contact_field):
        self.contactdata = {
            "fullname":         "A" + str(contact_row),
            "nickname":         "B" + str(contact_row),
            "company":          "C" + str(contact_row),
            "jobtitle":         "D" + str(contact_row),
            "picture":          "E" + str(contact_row),
            "phone1_label":     "F" + str(contact_row),
            "phone1":           "F" + str(contact_row + 1),
            "phone2_label":     "G" + str(contact_row),
            "phone2":           "G" + str(contact_row + 1),
            "phone3_label":     "H" + str(contact_row),
            "phone3":           "H" + str(contact_row + 1),
            "phone4_label":     "I" + str(contact_row),
            "phone4":           "I" + str(contact_row + 1),
            "email1_label":     "J" + str(contact_row),
            "email1":           "J" + str(contact_row + 1),
            "email2_label":     "K" + str(contact_row),
            "email2":           "K" + str(contact_row + 1),
            "sns1_label":       "L" + str(contact_row),
            "sns1":             "L" + str(contact_row + 1),
            "sns2_label":       "M" + str(contact_row),
            "sns2":             "M" + str(contact_row + 1),
            "sns3_label":       "N" + str(contact_row),
            "sns3":             "N" + str(contact_row + 1), 
            "address1_label":   "O" + str(contact_row),
            "address1":         "O" + str(contact_row + 1),
            "address2_label":   "P" + str(contact_row),
            "address2":         "P" + str(contact_row + 1),
            "year":             "Q" + str(contact_row),
            "month":            "R" + str(contact_row),
            "day":              "S" + str(contact_row),
            "id":               "T" + str(contact_row),
            "idcard_label":     "U" + str(contact_row),
            "idcard":           "U" + str(contact_row + 1),
            "bank_label":       "V" + str(contact_row),
            "bank":             "V" + str(contact_row + 1),
            "bank_branch":      "W" + str(contact_row),
            "note":             "Y" + str(contact_row),
            "photo":            "Z" + str(contact_row),
        }
        self.key = self.contactdata[contact_field]
        return self.worksheet[self.contactdata[contact_field]].value

    def getContactRowByID(self, id):
        for row in self.worksheet.iter_rows(min_row=6, min_col=20, max_col=20, values_only=False):
            for cell in row:
                if cell.value == id:
                    return cell.row


    def setContactData(self, contact_row, contact_field, contact_newdata):
        self.contactdata = {
            "fullname":         "A" + str(contact_row),
            "nickname":         "B" + str(contact_row),
            "company":          "C" + str(contact_row),
            "jobtitle":         "D" + str(contact_row),
            "picture":          "E" + str(contact_row),
            "phone1_label":     "F" + str(contact_row),
            "phone1":           "F" + str(contact_row + 1),
            "phone2_label":     "G" + str(contact_row),
            "phone2":           "G" + str(contact_row + 1),
            "phone3_label":     "H" + str(contact_row),
            "phone3":           "H" + str(contact_row + 1),
            "phone4_label":     "I" + str(contact_row),
            "phone4":           "I" + str(contact_row + 1),
            "email1_label":     "J" + str(contact_row),
            "email1":           "J" + str(contact_row + 1),
            "email2_label":     "K" + str(contact_row),
            "email2":           "K" + str(contact_row + 1),
            "sns1_label":       "L" + str(contact_row),
            "sns1":             "L" + str(contact_row + 1),
            "sns2_label":       "M" + str(contact_row),
            "sns2":             "M" + str(contact_row + 1),
            "sns3_label":       "N" + str(contact_row),
            "sns3":             "N" + str(contact_row + 1), 
            "address1_label":   "O" + str(contact_row),
            "address1":         "O" + str(contact_row + 1),
            "address2_label":   "P" + str(contact_row),
            "address2":         "P" + str(contact_row + 1),
            "year":             "Q" + str(contact_row),
            "month":            "R" + str(contact_row),
            "day":              "S" + str(contact_row),
            "id":               "T" + str(contact_row),
            "idcard_label":     "U" + str(contact_row),
            "idcard":           "U" + str(contact_row + 1),
            "bank_label":       "V" + str(contact_row),
            "bank":             "V" + str(contact_row + 1),
            "bank_branch":      "W" + str(contact_row),
            "note":             "Y" + str(contact_row),
            "photo":            "Z" + str(contact_row),
        }
        self.key = self.contactdata[contact_field]
        self.worksheet[self.contactdata[contact_field]].value = contact_newdata

    def removeContactData(self, row):
        self.worksheet.delete_rows(row, 2)
            

    def searchContactData(self, list, keyword):
        self.picPath = "./data/{0}/profile_pictures/".format(self.username)
        self.current_row = 6
        for row in self.worksheet.iter_rows(min_row=6, min_col=1, max_col=1, values_only=False):
            for cell in row:
                # Check if cell isn't empty
                if cell.value != None and str.lower(keyword) in str.lower(cell.value):
                    # Create custom listWidget
                    customListWidget = contakuto_basegui.QListWidgetX()
                    # Get value
                    nicknameValue = self.worksheet.cell(row=cell.row, column=2).value
                    companyValue = self.worksheet.cell(row=cell.row, column=3).value
                    picValue = self.worksheet.cell(row=cell.row, column=5).value
                    itemID = self.worksheet.cell(row=cell.row, column=20).value
                    # Set item icon
                    if not picValue:
                        iconValue = "./image/default_avatar.png"
                    else:
                        iconValue = self.picPath + picValue
                    iconPixmap = QPixmap(iconValue)
                    iconPixmap = iconPixmap.scaled(
                        40, 40,
                        aspectRatioMode=Qt.IgnoreAspectRatio, transformMode=Qt.SmoothTransformation)
                    customListWidget.setID(itemID)
                    customListWidget.setTextUp(cell.value)
                    if not nicknameValue and not companyValue:
                        pass
                    elif nicknameValue and not companyValue:
                        customListWidget.setTextDown(nicknameValue)
                    elif not nicknameValue and companyValue:
                        customListWidget.setTextDown(companyValue)
                    else:
                        customListWidget.setTextDown(nicknameValue + ", " + companyValue)
                    customListWidget.setIcon(iconPixmap)
                    # Add cell value to list
                    # Create item
                    item = QListWidgetItem(cell.value)
                    item.setSizeHint(customListWidget.sizeHint())
                    list.addItem(item)
                    list.setItemWidget(item, customListWidget)
                    #list.sortItems(Qt.AscendingOrder)
                self.current_row += 1

    def saveData(self):
        self.workbook.save(self.filepath)