from data_class import UserProfileData
import sys
import os
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
import openpyxl as xl
import contakuto_stylesheet as style
import contakuto_main as main
from contakuto_basegui import *

class LoginUI(QDialog):
    def __init__(self, parent=None):
        super().__init__()
        self.setWindowTitle("Contakuto")
        self.resize(765, 450)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setWindowFlags(Qt.FramelessWindowHint)
        self.setStyleSheet(style.css)

        self.background = QWidget(self)
        self.background.setObjectName("background")
        self.background.setGeometry(QRect(
            20,
            20,
            self.width() - 40,
            self.height() - 40))
        self.shadow = QGraphicsDropShadowEffect()
        self.shadow.setBlurRadius(10)
        self.shadow.setOffset(0, 0)
        self.shadow.setColor(QColor(0, 0, 0, 180))
        self.background.setGraphicsEffect(self.shadow)

        self.titlebar = QWidget(self)
        self.titlebar.setObjectName("titlebar")
        self.titlebar.setGeometry(QRect(
            self.background.x(), self.background.y(), self.background.width(), 40))

        self.exit_button = QPushButton(self.titlebar)
        self.exit_button.setObjectName("exit_button")
        self.exit_button.setCursor(QCursor(Qt.PointingHandCursor))
        self.exit_button.setFixedSize(16, 16)
        self.exit_button.move(17, 12)
        self.exit_button.clicked.connect(self.closeDialog)

        self.minimize_button = QPushButton(self.titlebar)
        self.minimize_button.setObjectName("minimize_button")
        self.minimize_button.setCursor(self.exit_button.cursor())
        self.minimize_button.setFixedSize(16, 16)
        self.minimize_button.move(
            self.exit_button.x() + self.exit_button.width() + 10, 
            self.exit_button.y())
        self.minimize_button.clicked.connect(self.showMinimized)

        self.fullscreen_button = QPushButton(self.titlebar)
        self.fullscreen_button.setObjectName("fullscreen_button")
        self.fullscreen_button.setCursor(self.exit_button.cursor())
        self.fullscreen_button.setFixedSize(16, 16)
        self.fullscreen_button.move(
            self.minimize_button.x() + self.minimize_button.width() + 10, 
            self.minimize_button.y())

        self.product_image = QLabel(self.background)
        self.product_image.setObjectName("product_image")
        self.product_image.setFixedSize(self.background.height(), self.background.height())
        self.product_image.move(
            self.background.width() - self.product_image.width(), 0)
        self.product_image_shadow = QGraphicsDropShadowEffect()
        self.product_image_shadow.setBlurRadius(10)
        self.product_image_shadow.setOffset(0)
        self.product_image.setGraphicsEffect(self.product_image_shadow)


        # Login widget
        self.login_box = QWidget(self.background)
        self.login_box.setObjectName("login_box")
        self.login_box.setFixedSize(
            self.background.width() - self.product_image.width(),
            self.background.height() - self.titlebar.height())
        self.login_box.move(0, self.titlebar.height())


        self.main_label = QLabel(self.login_box)
        self.main_label.setObjectName("main_label")
        self.main_label.setGeometry(QRect(
            0,
            50,
            self.background.width() - self.product_image.width(),
            40))
        self.main_label.setAlignment(Qt.AlignCenter | Qt.AlignTop)
        self.main_label.setStyleSheet("")
        self.main_label.setText("LOGIN")

        self.username = QLineEdit(self.login_box)
        self.username.setMaxLength(22)
        self.username.setFocus()
        self.username.setFixedSize(250, 40)
        self.username.move(
            (self.main_label.width() - self.username.width()) / 2,
            self.main_label.y() + 50)
        self.username.setPlaceholderText("Username")

        self.password = QLineEdit(self.login_box)
        self.password.setMaxLength(14)
        self.password.setFixedSize(self.username.size())
        self.password.move(
            self.username.x(),
            self.username.y() + self.username.height() + 10)
        self.password.setEchoMode(QLineEdit.Password)
        self.password.setPlaceholderText("Password")
        self.password.setStyleSheet(self.username.styleSheet())

        self.login_button = QPushButton(self.login_box)
        self.login_button.setObjectName("main_button")
        self.login_button.setDefault(True)
        self.login_button.setFixedSize(150, 30)
        self.login_button.move(
            (self.main_label.width() - self.login_button.width()) / 2,
            self.password.y() + self.password.height() + 20)
        self.login_button.setText("Login")
        self.login_button.setCursor(QCursor(Qt.PointingHandCursor))
        self.login_button_shadow = QGraphicsDropShadowEffect()
        self.login_button_shadow.setBlurRadius(15)
        self.login_button_shadow.setOffset(0)
        self.login_button.setGraphicsEffect(self.login_button_shadow)
        self.login_button.clicked.connect(self.login)

        self.forgotpw = QPushButton(self.login_box)
        self.forgotpw.setObjectName("forgotpw-register-login")
        self.forgotpw.setFixedSize(200, 25)
        self.forgotpw.move(
            (self.main_label.width() - self.forgotpw.width()) / 2,
            self.login_button.y() + self.login_button.height() + 25)
        self.forgotpw.setText("Forgot password?")
        self.forgotpw.setCursor(QCursor(Qt.PointingHandCursor))
        self.forgotpw.clicked.connect(self.forgotpwDialog)

        self.register = QPushButton(self.login_box)
        self.register.setObjectName(self.forgotpw.objectName())
        self.register.setFixedSize(200, 25)
        self.register.setCursor(self.forgotpw.cursor())
        self.register.move(
            (self.main_label.width() - self.register.width()) / 2,
            self.forgotpw.y() + self.forgotpw.height() + 7)
        self.register.setText("Create new account")
        self.register.clicked.connect(self.registerDialog)

        # Registering widget
        self.register_box = QWidget(self.background)
        self.register_box.setObjectName("register_box")
        self.register_box.setFixedSize(self.login_box.width(), self.login_box.height())
        self.register_box.move(self.login_box.pos())

        self.register_label = QLabel(self.register_box)
        self.register_label.setObjectName("main_label")
        self.register_label.setAlignment(self.main_label.alignment())
        self.register_label.setFixedSize(self.main_label.size())
        self.register_label.move(0, 10)
        self.register_label.setText("REGISTER: Step 1")

        self.register_fullname = QLineEdit(self.register_box)
        self.register_fullname.setPlaceholderText("Full Name")
        self.register_fullname.setMaxLength(22)
        self.register_fullname.setGeometry(
            self.username.x(),
            self.register_label.y() + self.register_label.height() + 30,
            self.username.width(),
            self.username.height())

        self.register_username = QLineEdit(self.register_box)
        self.register_username.setPlaceholderText("Username")
        self.register_username.setValidator(QRegExpValidator(QRegExp("^[a-zA-Z0-9_]{6,14}$")))
        self.register_username.setGeometry(
            self.register_fullname.x(),
            self.register_fullname.y() + self.register_fullname.height() + 10,
            self.register_fullname.width(),
            self.register_fullname.height())

        self.register_language = QComboBox(self.register_box)
        self.register_language.addItems(("English", "Tiếng Việt", "日本語"))
        self.register_language.setCursor(QCursor(Qt.PointingHandCursor))
        self.register_language.setFixedSize(130, 40)
        self.register_language.move(
            (self.main_label.width() - self.register_language.width()) / 2,
            self.register_username.y() + self.register_username.height() + 10)
            
        self.register_nextstep = QPushButton(self.register_box)
        self.register_nextstep.setObjectName("main_button")
        self.register_nextstep.setCursor(self.login_button.cursor())
        self.register_nextstep.setText("Next step")
        self.register_nextstep.setFixedSize(self.login_button.size())
        self.register_nextstep.move(
            self.login_button.x(),
            self.register_language.y() + self.register_language.height() + 20)
        self.register_ns_shadow = QGraphicsDropShadowEffect()
        self.register_ns_shadow.setBlurRadius(self.login_button_shadow.blurRadius())
        self.register_ns_shadow.setOffset(self.login_button_shadow.offset())
        self.register_nextstep.setGraphicsEffect(self.register_ns_shadow)
        self.register_nextstep.clicked.connect(self.registerNextStep)

        self.register_login_button = QPushButton(self.register_box)
        self.register_login_button.setObjectName(self.forgotpw.objectName())
        self.register_login_button.setCursor(self.forgotpw.cursor())
        self.register_login_button.setText("← Already have an account?")
        self.register_login_button.setFixedSize(200, self.forgotpw.height())
        self.register_login_button.move(
            5,
            self.login_box.height() - self.register_login_button.height() - 10)
        self.register_login_button.clicked.connect(self.loginDialog)

        # Registering widget 2
        self.register_box2 = QWidget(self.background)
        self.register_box2.setObjectName("register_box")
        self.register_box2.setFixedSize(self.login_box.width(), self.login_box.height())
        self.register_box2.move(self.login_box.pos())
        
        self.register_label2 = QLabel(self.register_box2)
        self.register_label2.setObjectName("main_label")
        self.register_label2.setAlignment(self.main_label.alignment())
        self.register_label2.setFixedSize(self.main_label.size())
        self.register_label2.move(0, 10)
        self.register_label2.setText("REGISTER: Step 2")
        
        self.register_password = QLineEdit(self.register_box2)
        self.register_password.setPlaceholderText("Password")
        self.register_password.setEchoMode(QLineEdit.Password)
        self.register_password.setValidator(QRegExpValidator(QRegExp("^[a-zA-Z0-9_]{6,14}$")))
        self.register_password.setGeometry(QRect(
            self.register_fullname.x(),
            self.register_label2.y() + self.register_label.height() + 30,
            self.register_fullname.width(),
            self.register_fullname.height()))

        self.register_password_cf = QLineEdit(self.register_box2)
        self.register_password_cf.setPlaceholderText("Confirm password")
        self.register_password_cf.setEchoMode(QLineEdit.Password)
        self.register_password_cf.setValidator(QRegExpValidator(QRegExp("^[a-zA-Z0-9_]{6,14}$")))
        self.register_password_cf.setGeometry(QRect(
            self.register_fullname.x(),
            self.register_password.y() + self.register_password.height() + 10,
            self.register_fullname.width(),
            self.register_fullname.height()))
        
        self.register_pin = QLineEdit(self.register_box2)
        self.register_pin.setPlaceholderText("PIN")
        self.register_pin.setEchoMode(QLineEdit.Password)
        self.register_pin.setValidator(QRegExpValidator(QRegExp("^[0-9]{4,4}$")))
        self.register_pin.setGeometry(QRect(
            self.register_fullname.x(),
            self.register_password_cf.y() + self.register_password_cf.height() + 10,
            self.register_fullname.width(),
            self.register_fullname.height()))

        self.register_button = QPushButton(self.register_box2)
        self.register_button.setObjectName("main_button")
        self.register_button.setCursor(self.login_button.cursor())
        self.register_button.setText("Register")
        self.register_button.setFixedSize(self.login_button.size())
        self.register_button.move(
            self.login_button.x(),
            self.register_pin.y() + self.register_pin.height() + 20)
        self.register_btn_shadow = QGraphicsDropShadowEffect()
        self.register_btn_shadow.setBlurRadius(self.login_button_shadow.blurRadius())
        self.register_btn_shadow.setOffset(self.login_button_shadow.offset())
        self.register_button.setGraphicsEffect(self.register_btn_shadow)
        self.register_button.clicked.connect(self.registerFinalStep)

        self.register_login_button2 = QPushButton(self.register_box2)
        self.register_login_button2.setObjectName(self.forgotpw.objectName())
        self.register_login_button2.setCursor(self.forgotpw.cursor())
        self.register_login_button2.setText("← Already have an account?")
        self.register_login_button2.setFixedHeight(self.forgotpw.height())
        self.register_login_button2.move(
            15,
            self.login_box.height() - self.register_login_button.height() - 10)
        self.register_login_button2.clicked.connect(self.loginDialog)

        self.register_backstep = QPushButton(self.register_box2)
        self.register_backstep.setObjectName(self.forgotpw.objectName())
        self.register_backstep.setCursor(self.forgotpw.cursor())
        self.register_backstep.setText("← Back to step 1")
        self.register_backstep.setFixedHeight(self.forgotpw.height())
        self.register_backstep.move(
            self.register_login_button2.x(),
            self.register_login_button2.y() - self.register_backstep.height())
        self.register_backstep.clicked.connect(self.backStep)

        # Forgot password widget
        self.forgotpw_box = QWidget(self.background)
        self.forgotpw_box.setObjectName("forgotpw_box")
        self.forgotpw_box.setFixedSize(self.login_box.size())
        self.forgotpw_box.move(self.login_box.pos())
        
        self.forgotpw_label = QLabel(self.forgotpw_box)
        self.forgotpw_label.setObjectName("main_label")
        self.forgotpw_label.setAlignment(self.main_label.alignment())
        self.forgotpw_label.setFixedSize(self.main_label.size())
        self.forgotpw_label.move(0, 10)
        self.forgotpw_label.setText("Reset password")

        self.forgotpw_username = QLineEdit(self.forgotpw_box)
        self.forgotpw_username.setPlaceholderText("Username")
        self.forgotpw_username.setGeometry(QRect(
            self.username.x(),
            self.forgotpw_label.y() + self.forgotpw_label.height() + 10,
            self.username.width(),
            self.username.height()))
        self.forgotpw_username.setValidator(QRegExpValidator(QRegExp("^[a-zA-Z0-9_]{6,14}$")))
        
        self.forgotpw_pin = QLineEdit(self.forgotpw_box)
        self.forgotpw_pin.setPlaceholderText("PIN")
        self.forgotpw_pin.setGeometry(QRect(
            self.username.x(),
            self.forgotpw_username.y() + self.forgotpw_username.height() + 10,
            self.username.width(),
            self.username.height()))
        self.forgotpw_pin.setValidator(QRegExpValidator(QRegExp("^[0-9_]{4,4}$")))

        self.forgotpw_newpw = QLineEdit(self.forgotpw_box)
        self.forgotpw_newpw.setPlaceholderText("New password")
        self.forgotpw_newpw.setFixedSize(self.forgotpw_username.size())
        self.forgotpw_newpw.move(
            self.forgotpw_username.x(),
            self.forgotpw_pin.y() + self.forgotpw_pin.height() + 10)
        self.forgotpw_newpw.setValidator(QRegExpValidator(QRegExp("^[a-zA-Z0-9_]{6,14}$")))

        self.forgotpw_newpw_cf = QLineEdit(self.forgotpw_box)
        self.forgotpw_newpw_cf.setPlaceholderText("Confirm password")
        self.forgotpw_newpw_cf.setFixedSize(self.forgotpw_username.size())
        self.forgotpw_newpw_cf.move(
            self.username.x(),
            self.forgotpw_newpw.y() + self.forgotpw_newpw.height() + 10)
        self.forgotpw_newpw_cf.setValidator(QRegExpValidator(QRegExp("^[a-zA-Z0-9_]{6,14}$")))

        self.forgotpw_reset = QPushButton(self.forgotpw_box)
        self.forgotpw_reset.setObjectName("main_button")
        self.forgotpw_reset.setCursor(self.login_button.cursor())
        self.forgotpw_reset.setText("Reset")
        self.forgotpw_reset.setFixedSize(self.login_button.size())
        self.forgotpw_reset.move(
            self.login_button.x(),
            self.forgotpw_newpw_cf.y() + self.forgotpw_newpw_cf.height() + 20)
        self.forgotpw_reset_shadow = QGraphicsDropShadowEffect()
        self.forgotpw_reset_shadow.setBlurRadius(self.login_button_shadow.blurRadius())
        self.forgotpw_reset_shadow.setOffset(self.login_button_shadow.offset())
        self.forgotpw_reset.setGraphicsEffect(self.forgotpw_reset_shadow)

        self.forgotpw_login = QPushButton(self.forgotpw_box)
        self.forgotpw_login.setObjectName("forgotpw-register-login")
        self.forgotpw_login.setCursor(QCursor(Qt.PointingHandCursor))
        self.forgotpw_login.setText("← Login")
        self.forgotpw_login.setFixedSize(90, self.forgotpw.height())
        self.forgotpw_login.move(
            5,
            self.login_box.height() - self.register_login_button.height() - 10)
        self.forgotpw_login.clicked.connect(self.loginDialog)



        self.oldPos = self.pos()
        self.login_box.raise_()
        self.product_image.raise_()

    def mousePressEvent(self, event):
        self.oldPos = event.globalPos()

    def mouseMoveEvent(self, event):
        #time.sleep(0.02)  # sleep for 20ms
        delta = QPoint(event.globalPos() - self.oldPos)
        self.move(self.x() + delta.x(), self.y() + delta.y())
        self.oldPos = event.globalPos()

    def closeDialog(self):
        sys.exit()

    def loginDialog(self):
        self.login_button.setDefault(True)
        self.register_button.setDefault(False)
        self.forgotpw_reset.setDefault(False)
        self.register_fullname.clear()
        self.register_username.clear()
        self.register_password.clear()
        self.register_pin.clear()
        self.forgotpw_username.clear()
        self.forgotpw_newpw.clear()
        self.forgotpw_newpw_cf.clear()
        self.forgotpw_pin.clear()
        self.login_box.raise_()
        self.product_image.raise_()

    def registerDialog(self):
        self.login_button.setDefault(False)
        self.register_button.setDefault(True)
        self.forgotpw_reset.setDefault(False)
        self.username.clear()
        self.password.clear()
        self.forgotpw_username.clear()
        self.forgotpw_newpw.clear()
        self.forgotpw_newpw_cf.clear()
        self.forgotpw_pin.clear()
        self.register_box.raise_()
        self.product_image.raise_() 

    def forgotpwDialog(self):
        self.login_button.setDefault(False)
        self.register_button.setDefault(False)
        self.forgotpw_reset.setDefault(True)
        self.username.clear()
        self.password.clear()
        self.register_username.clear()
        self.register_password.clear()
        self.register_password_cf.clear()
        self.register_pin.clear()
        self.forgotpw_box.raise_()
        self.product_image.raise_()
    
    def backStep(self):
        self.register_box.raise_()
        self.product_image.raise_()

    def login(self):
        # Open userdata file. If file exists, then check password
        try:
            self.workbook = xl.load_workbook("./data/{0}/{0}.xlsx".format(self.username.text()))
            self.worksheet = self.workbook["Sheet"]
            self.password_data = self.worksheet["C2"].value

            # If password is wrong
            if self.password_data != self.password.text():
                self.message = QMessageBoxX(
                    icon = "warning", 
                    boldtext = "Wrong password!",
                    text = "Your password seems to be wrong. Please try again!",
                    ok = True,
                    cancel = False)
                self.message.exec()
                self.password.clear()
            # If password is right, then login successfully
            else:
                self.message = QMessageBoxX(
                    icon = "information",
                    boldtext = "Login successfully",
                    text = "You have logged in successfully as [{0}]".format(self.username.text()),
                    ok = True,
                    cancel = False)
                self.message.exec()
                main.user = UserProfileData(self.username.text())
                main.theme = "light"
                main.mainui = main.MainUI()
                main.mainui.show()
                self.close()

        except FileNotFoundError:
            self.message = QMessageBoxX(
                icon = "warning",
                boldtext = "User doesn't exist",
                text = "This username doesn't exist. Please check again!",
                ok = True,
                cancel = False)
            self.message.exec()
            self.password.clear()
            
    def registerNextStep(self):
        if len(self.register_fullname.text()) < 4:
            warning = QMessageBoxX(
                icon = "warning",
                boldtext = "Too short full name",
                text = "Full name must contain at least 4 characters.")
            warning.exec()
        elif len(self.register_username.text()) < 6:
            warning = QMessageBoxX(
                icon = "warning",
                boldtext = "Too short username",
                text = "Username must contain at least 6 characters.")
            warning.exec()
        else:
            self.register_box2.raise_()
            self.product_image.raise_()

    def registerFinalStep(self):
        if len(self.register_password.text()) < 6:
            warning = QMessageBoxX(
                icon = "warning",
                boldtext = "Too short password",
                text = "Password must contain at least 6 characters.")
            warning.exec()
        elif self.register_password_cf.text() != self.register_password.text():
            warning = QMessageBoxX(
                icon = "warning",
                boldtext = "Wrong confirmation password",
                text = "The confirmation password you have entered is wrong. Please check again.")
        elif len(self.register_pin.text()) < 4:
            warning = QMessageBox(
                icon = "warning",
                boldtext = " Too short PIN",
                text = "PIN must contain at least 4 numbers")
            warning.exec()
        else:
            try:
                self.workbook = xl.load_workbook("./data/{0}/{0}.xlsx".format(self.register_username.text()))
                self.warning = QMessageBoxX(
                    icon = "warning",
                    boldtext = "Existed username",
                    text = "This username has been used. Please choose another username.")
            # If inputed username doesn't exist, write all registered data
            except FileNotFoundError:
                self.workbook = xl.Workbook()
                self.worksheet = self.workbook.active
                # Create user fields' labels
                self.userdatalabel = iter([
                                            "Fullname",
                                            "Username",
                                            "Password",
                                            "PIN",
                                            "Language",
                                            "Profile Picture",
                                            "Theme"
                                        ])
                self.cell_range = self.worksheet["A1:G1"]
                for row in self.cell_range:
                    for cell in row:
                        cell.value = next(self.userdatalabel)
                # Check language combobox value
                if self.register_language.currentIndex() == 0:
                    self.regLanguage = "english"
                elif self.register_language.currentIndex() == 1:
                    self.regLanguage = "vietnamese"
                elif self.register_language.currentIndex() == 2:
                    self.regLanguage = "japanese"
                self.userdata = iter([
                                        self.register_fullname.text(),
                                        self.register_username.text(),
                                        self.register_password.text(),
                                        self.register_pin.text(),
                                        self.regLanguage,
                                        None,
                                        "light"
                                    ])
                # Set user data
                self.cell_range = self.worksheet["A2:G2"]
                for row in self.cell_range:
                    for cell in row:
                        cell.value = next(self.userdata)
                # Create contacts fields' label
                self.usercontactlabel = iter([
                                                "Fullname",
                                                "Nickname",
                                                "Company",
                                                "Job Title",
                                                "Picture",
                                                "Phone1",
                                                "Phone2",
                                                "Phone3",
                                                "Phone4",
                                                "Email1",
                                                "Email2",
                                                "SNS1",
                                                "SNS2",
                                                "SNS3",
                                                "Address1",
                                                "Address2",
                                                "Year",
                                                "Month",
                                                "Day",
                                                "ID",
                                                "ID Card",
                                                "Bank",
                                                "Bank Branch",
                                                "Note",
                                                "Photos",
                                            ])
                self.cell_range = self.worksheet["A5:Y5"]
                for row in self.cell_range:
                    for cell in row:
                        cell.value = next(self.usercontactlabel)
                # Create user data's folder
                os.makedirs("./data/{0}".format(self.register_username.text()))
                os.makedirs("./data/{0}/profile_pictures".format(self.register_username.text()))
                # Save user data's file
                self.workbook.save("./data/{0}/{0}.xlsx".format(self.register_username.text()))
                self.information = QMessageBoxX(
                    icon = "information",
                    boldtext = "Registered successfully",
                    text = "Your account [{0}] has been created successfully!".format(
                        self.register_username.text()))
                self.information.exec()
                self.loginDialog()
        
        

    
if __name__ == "__main__":
    app = QApplication(sys.argv)
    QFontDatabase.addApplicationFont("./image/fonts/SanFranciscoDisplay-Regular.otf")
    QFontDatabase.addApplicationFont("./image/fonts/SanFranciscoDisplay-Bold.otf")
    QFontDatabase.addApplicationFont("./image/fonts/SanFranciscoDisplay-Medium.otf")
    QFontDatabase.addApplicationFont("./image/fonts/SanFranciscoDisplay-Thin.otf")
    mainui = MainUI()
    mainui.show()
    sys.exit(app.exec())



